import os
from typing import List, Dict, Any, Optional, Union
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelUtil:

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self._load()

    def _load(self):
        """ 加载 Excel 文件，如果文件不存在则自动创建 """
        if not os.path.exists(self.file_path):
            wb = Workbook()
            wb.save(self.file_path)

        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook.active

    def _get_sheet(self, sheet_name: Optional[str]):
        """
        获取指定工作表
        :param sheet_name: 工作表名称（None表示当前sheet）
        :return: Worksheet对象
        """
        if sheet_name:
            return self.workbook[sheet_name]
        return self.sheet

    # ========================
    # Sheet 操作
    # ========================
    def list_sheets(self) -> List[str]:
        """
        获取所有工作表名称
        :return: 工作表名称列表
        """
        return self.workbook.sheetnames

    def select_sheet(self, sheet_name: str):
        """
        选择当前工作表
        :param sheet_name: 工作表名称
        :return:
        """
        self.sheet = self.workbook[sheet_name]

    def add_sheet(self, sheet_name: str):
        """
        添加新的工作表
        :param sheet_name: 工作表名称
        :return: 是否创建成功
        """
        if sheet_name in self.workbook.sheetnames:
            return False
        self.sheet = self.workbook.create_sheet(sheet_name)
        return True

    # ========================
    # 单元格
    # ========================
    def read_cell(self, cell: str, sheet_name=None):
        """
        读取单元格内容
        :param cell: 单元格位置（如 A1）
        :param sheet_name: 工作表名称
        :return: 单元格值
        """
        sheet = self._get_sheet(sheet_name)
        return sheet[cell].value

    def write_cell(self, cell: str, value, sheet_name=None):
        """
        写入单元格内容
        :param cell: 单元格位置（如 A1）
        :param value: 写入的值
        :param sheet_name: 工作表名称
        :return:
        """
        sheet = self._get_sheet(sheet_name)
        sheet[cell].value = value

    # ========================
    # 行操作
    # ========================
    def read_row(self, row: int, sheet_name=None) -> List[Any]:
        """
        读取整行数据
        :param row: 行号（从1开始）
        :param sheet_name: 工作表名称
        :return: 行数据列表
        """
        sheet = self._get_sheet(sheet_name)
        return [cell.value for cell in sheet[row]]

    def write_row(self, row: int, data: List[Any], sheet_name=None):
        """
        写入整行数据
        :param row: 行号（从1开始）
        :param data: 要写入的数据列表
        :param sheet_name: 工作表名称
        :return:
        """
        sheet = self._get_sheet(sheet_name)
        for col, value in enumerate(data, start=1):
            sheet.cell(row=row, column=col).value = value

    # ========================
    # 列操作
    # ========================
    def read_column(self, col: Union[int, str], sheet_name=None) -> List[Any]:
        """
        读取整列数据
        :param col: 列索引或列字母（如 1 或 'A'）
        :param sheet_name: 工作表名称
        :return: 列数据列表
        """
        sheet = self._get_sheet(sheet_name)

        if isinstance(col, str):
            col = column_index_from_string(col)

        return [sheet.cell(row=r, column=col).value for r in range(1, sheet.max_row + 1)]

    def write_column(self, col: Union[int, str], data: List[Any], sheet_name=None):
        """
        写入整列数据
        :param col: 列索引或列字母（如 1 或 'A'）
        :param data: 要写入的数据列表
        :param sheet_name: 工作表名称
        :return:
        """
        sheet = self._get_sheet(sheet_name)

        if isinstance(col, str):
            col = column_index_from_string(col)

        for row, value in enumerate(data, start=1):
            sheet.cell(row=row, column=col).value = value

    # ========================
    # 区域
    # ========================
    def read_range(self, start: str, end: str, sheet_name=None):
        """
        读取指定区域数据
        :param start: 起始单元格（如 A1）
        :param end: 结束单元格（如 C3）
        :param sheet_name: 工作表名称
        :return: 二维数组数据
        """
        sheet = self._get_sheet(sheet_name)
        return [[cell.value for cell in row] for row in sheet[start:end]]

    # ========================
    # 保存
    # ========================
    def save(self):
        """ 保存 Excel 文件 """
        self.workbook.save(self.file_path)

    def close(self):
        """ 关闭工作簿 """
        self.workbook.close()